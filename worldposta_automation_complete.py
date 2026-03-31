#!/usr/bin/env python3
"""
WorldPosta Precision Link & CTA Checker
========================================
Uses sitemap + CTA-BUTTON-MAPPING to test every page and button.
Browser engine: Selenium + undetected_chromedriver (same as registration script).

Outputs a colour-coded Excel report with 4 sheets:
  1. Page Health     — every sitemap URL + HTTP status
  2. CTA Buttons     — every mapped button, expected vs actual destination
  3. ⚠ Issues Only  — filtered view of all problems
  4. 📊 Summary      — totals at a glance

Usage (GitHub Actions / n8n):
    python worldposta_automation_complete.py --headless
    python worldposta_automation_complete.py --headless --no-fail
"""

import os, re, sys, time, json, argparse, subprocess, requests
from urllib.parse import urljoin, urlparse
from datetime import datetime

# Selenium — same stack as registration script
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

BASE_URL        = "https://www.worldposta.com"
REQUEST_TIMEOUT = 15
PAGE_LOAD_TIMEOUT = 60          # seconds (same as registration script)
DEFAULT_TIMEOUT   = 30
SCREENSHOT_DIR  = "screenshots"
EXCEL_OUTPUT    = "link_check_report.xlsx"
JSON_OUTPUT     = "link_check_results.json"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}

# ─────────────────────────────────────────────────────────────────────────────
# ALL PAGES  (from sitemap.xml)
# ─────────────────────────────────────────────────────────────────────────────

SITEMAP_PAGES = [
    # Home
    ("Home",                             "/"),
    # CloudEdge
    ("CloudEdge Main",                   "/cloudedge"),
    ("Cloud Plan",                       "/CloudPlan"),
    ("Cloud Resource Optimization",      "/cloud-resource-optimization"),
    ("Dedicated Compute Instances",      "/dedicated-compute-instances"),
    ("Cloud Ransomware Protection",      "/cloud-ransomware-protection"),
    ("Comprehensive Cloud Security",     "/comprehensive-cloud-security-solutions"),
    ("VDC vs Traditional VMs",           "/virtual-data-centre-vs-traditional-vms"),
    ("Virtual Data Center Solutions",    "/virtual-data-center-solutions"),
    ("GPU Instances",                    "/gpu-instances"),
    ("Cloud Firewall",                   "/cloud-firewall"),
    # Posta
    ("Posta Main",                       "/posta"),
    ("Office 365 with Posta",            "/office-365-with-posta"),
    ("Google Business Email with Posta", "/google-business-email-with-posta"),
    ("Posta Hybrid",                     "/posta-hybrid"),
    ("Posta Gate",                       "/posta-gate"),
    ("Posta Gate Comparison",            "/posta-gate-comparison"),
    ("Posta Gate Security",              "/posta-gate-security"),
    ("Email Security Test",              "/email-security-test"),
    # CloudSpace
    ("CloudSpace Main",                  "/cloudspace"),
    # Pricing
    ("CloudEdge Pricing",                "/cloudedge-pricing"),
    ("Posta Pricing",                    "/posta-pricing"),
    # Solutions
    ("WPSYS IT Solutions",               "/wpsys-it-solutions"),
    ("IT Software Automation Tools",     "/it-software-automation-tools"),
    ("Cloud Based SOC Solutions",        "/cloud-based-soc-solutions"),
    ("Cloud Backup",                     "/cloud-backup"),
    ("Mission Critical Data Backup",     "/mission-critical-data-backup"),
    ("SAP Cloud Operations",             "/sap-cloud-operations"),
    ("Posta Portal",                     "/posta-portal"),
    ("CloudEdge Portal",                 "/cloudedge-portal"),
    ("Seamless Migration",               "/seamless-migration"),
    ("Cloud Security Solutions",         "/cloud-security-solutions"),
    ("Email Migration",                  "/email-migration"),
    ("Veeam Cloud Connect",              "/veeam-cloud-connect"),
    ("Cloud Solutions for SAP Partners", "/cloud-solutions-for-sap-partners"),
    # Support & Info
    ("Contact Us",                       "/contact-us"),
    ("Knowledge Base",                   "/knowledge-base"),
    ("Become a Partner",                 "/become-a-partner"),
    ("About WorldPosta",                 "/about-worldposta"),
    ("Terms",                            "/terms"),
    ("Free Trial",                       "/free-trial"),
    ("IT Excellence",                    "/It-excellence"),
    ("One SLA",                          "/one-sla"),
    ("Privacy Policy",                   "/privacy-policy"),
    ("Blogs",                            "/blogs"),
    ("E-Books",                          "/e-books"),
    # Standalone
    ("GPU Instance Pro",                 "/gpu-instance-pro"),
    ("UAE Home",                         "/uae-home"),
    ("Login Hub",                        "/login-hub"),
    # Support KB
    ("WP Support",                       "/wp-support"),
    ("CloudEdge Support",                "/cloudedge-support"),
    ("CloudSpace Support",               "/cloudspace-support"),
]

# ─────────────────────────────────────────────────────────────────────────────
# CTA BUTTON MAP  (from CTA-BUTTON-MAPPING.md)
#
#  (source_path, cta_text, expected_dest, css_hint, section)
#
#  expected_dest values:
#    "/path"          → internal page to HTTP-check
#    "https://..."    → external URL to HTTP-check
#    "sso-dynamic"    → auth-aware button, skip (can't test unauthenticated)
#    "form-submit"    → form submit button, no navigation target
#    "calendly-popup" → opens JS popup
#    "js-function"    → pure JS, no URL
# ─────────────────────────────────────────────────────────────────────────────

CTA_MAP = [
    # ── Home ──────────────────────────────────────────────────────────────────
    ("/",          "Get Started",                    "sso-dynamic",           "btn__primary btn__hi",                 "Hero slider"),
    ("/",          "Explore CloudEdge",              "/cloudedge",            "btn__icon btn__lg RegesterBtn",        "CloudEdge section"),
    ("/",          "Explore SAP Operations",         "/sap-cloud-operations", "btn__icon btn__lg RegesterBtn",        "SAP section"),
    ("/",          "Explore CloudSpace",             "/cloudspace",           "btn__icon btn__lg RegesterBtn",        "CloudSpace section"),
    ("/",          "Explore WPSYS IT solutions",     "/wpsys-it-solutions",   "btn__icon btn__lg RegesterBtn",        "WPSYS section"),
    ("/",          "Subscribe",                      "form-submit",           "subscribeStyleBtn",                    "Footer newsletter"),
    # ── CloudEdge ─────────────────────────────────────────────────────────────
    ("/cloudedge", "Sign Up",                        "sso-dynamic",           "btn__primary",                         "Hero"),
    # ── Comprehensive Cloud Security ──────────────────────────────────────────
    ("/comprehensive-cloud-security-solutions", "Contact Sales", "/contact-us", "btn__primary btn__hi",              "Bottom CTA"),
    # ── GPU Instances ──────────────────────────────────────────────────────────
    ("/gpu-instances", "Contact Sales",              "/contact-us",           "btn__primary btn__hi",                 "Multiple sections"),
    # ── VDC vs Traditional VMs ────────────────────────────────────────────────
    ("/virtual-data-centre-vs-traditional-vms", "Schedule a Technical Consultation", "calendly-popup", "btn__primary btn__hi", "CTA section"),
    ("/virtual-data-centre-vs-traditional-vms", "Contact Sales", "/contact-us", "btn__primary btn__hi",              "Bottom CTA"),
    # ── Cloud Firewall ────────────────────────────────────────────────────────
    ("/cloud-firewall", "Contact Sales",             "/contact-us",           "btn__primary btn__hi",                 "Bottom CTA"),
    # ── Posta Main ────────────────────────────────────────────────────────────
    ("/posta",     "Sign Up",                        "sso-dynamic",           "btn__primary btn__hi minwidth-170",    "Hero"),
    ("/posta",     "Read More",                      "/one-sla",              "btn__secondary btn__link",             "One-SLA section"),
    ("/posta",     "Read More",                      "/mission-critical-data-backup", "btn__secondary btn__link",    "Backup section"),
    # ── Office 365 with Posta ─────────────────────────────────────────────────
    ("/office-365-with-posta", "Contact Sales",      "/contact-us",           "btn__primary btn__hi",                 "Bottom section"),
    # ── Posta Gate Security ───────────────────────────────────────────────────
    ("/posta-gate-security", "Contact Sales",        "/contact-us",           "btn__primary btn__hi",                 "Bottom CTA"),
    # ── Email Security Test ───────────────────────────────────────────────────
    ("/email-security-test", "Test Your Email Security", "form-submit",       "btn__primary btn__hi",                 "Interactive form"),
    # ── CloudSpace ────────────────────────────────────────────────────────────
    ("/cloudspace", "Sign Up",                       "sso-dynamic",           "btn__primary btn__hi",                 "Hero"),
    # ── Posta Pricing ─────────────────────────────────────────────────────────
    ("/posta-pricing", "Get Started Now",            "sso-dynamic",           "btn__secondary btn__hi minwidth-170",  "Per-plan CTA"),
    # ── CloudEdge Pricing ─────────────────────────────────────────────────────
    ("/cloudedge-pricing", "Get Started",            "sso-dynamic",           "btn__primary btn__hi",                 "Per-plan CTA"),
    # ── Cloud Backup ──────────────────────────────────────────────────────────
    ("/cloud-backup", "Submit Request",              "form-submit",           "btn__xl btn__block bgBtnstyle",         "Backup inquiry form"),
    # ── Cloud Security Solutions ──────────────────────────────────────────────
    ("/cloud-security-solutions", "Get Started",     "/contact-us",           "btn__primary btn__hi",                 "Ready to Secure section"),
    # ── Posta Portal ──────────────────────────────────────────────────────────
    ("/posta-portal", "Contact Sales",               "/contact-us",           "btn__primary btn__hi minwidth-170",    "Bottom CTA"),
    ("/posta-portal", "Schedule a Demo",             "js-function",           "btn__secondary btn__hi",               "requestDemo()"),
    # ── CloudEdge Portal ──────────────────────────────────────────────────────
    ("/cloudedge-portal", "Contact Sales",           "/contact-us",           "btn__primary btn__hi minwidth-170",    "Bottom CTA"),
    # ── Seamless Migration ────────────────────────────────────────────────────
    ("/seamless-migration", "Contact Sales",         "/contact-us",           "btn__primary btn__hi",                 "Bottom CTA"),
    # ── Email Migration ───────────────────────────────────────────────────────
    ("/email-migration", "Contact Sales",            "/contact-us",           "btn__primary btn__hi",                 "Bottom CTA"),
    # ── Contact Us ────────────────────────────────────────────────────────────
    ("/contact-us", "Submit Request",                "form-submit",           "btn__secondary btn__block",             "Contact form"),
    ("/contact-us", "Privacy Policy and Terms",      "/terms",                "inline link",                          "Form checkbox label"),
    # ── Free Trial ────────────────────────────────────────────────────────────
    ("/free-trial", "Submit Request",                "form-submit",           "btn__xl btn__block bgBtnstyle",         "Free trial form"),
    # ── Become a Partner ──────────────────────────────────────────────────────
    ("/become-a-partner", "Submit Request",          "form-submit",           "btn__secondary btn__block",             "Partner inquiry form"),
    # ── IT Excellence ─────────────────────────────────────────────────────────
    ("/It-excellence", "Get Started Today",          "sso-dynamic",           "btn__primary",                         "Hero CTA"),
    # ── GPU Instance Pro ──────────────────────────────────────────────────────
    ("/gpu-instance-pro", "Start Your GPU Journey Today", "/contact-us",      "btn__primary btn__hi",                 "Bottom CTA"),
    # ── Login Hub ─────────────────────────────────────────────────────────────
    ("/login-hub", "Sign In",                        "js-function",           "brand green button",                   "onButtonClick()"),
    # ── UAE Home ──────────────────────────────────────────────────────────────
    ("/uae-home", "Get Started",                     "sso-dynamic",           "btn__primary btn__hi",                 "UAE hero"),
    # ── Header Nav ────────────────────────────────────────────────────────────
    ("nav", "CloudEdge",              "/cloudedge",              "nav link", "Products dropdown"),
    ("nav", "Posta",                  "/posta",                  "nav link", "Products dropdown"),
    ("nav", "Posta Gate",             "/posta-gate-security",    "nav link", "Products dropdown"),
    ("nav", "CloudSpace",             "/cloudspace",             "nav link", "Products dropdown"),
    ("nav", "Cloud Backup",           "/cloud-backup",           "nav link", "Solutions dropdown"),
    ("nav", "WPSYS IT Solutions",     "/wpsys-it-solutions",     "nav link", "Solutions dropdown"),
    ("nav", "SAP Cloud Operations",   "/sap-cloud-operations",   "nav link", "Solutions dropdown"),
    ("nav", "Posta Pricing",          "/posta-pricing",          "nav link", "Pricing dropdown"),
    ("nav", "CloudEdge Pricing",      "/cloudedge-pricing",      "nav link", "Pricing dropdown"),
    ("nav", "Contact Us",             "/contact-us",             "nav link", "Top-level nav"),
    ("nav", "Knowledge Base",         "/knowledge-base",         "nav link", "Support dropdown"),
    ("nav", "Blogs",                  "/blogs",                  "nav link", "Support dropdown"),
    ("nav", "About Us",               "/about-worldposta",       "nav link", "Support dropdown"),
    # ── Footer ────────────────────────────────────────────────────────────────
    ("footer", "About Us",            "/about-worldposta",       "footer link", "Company"),
    ("footer", "Become a Partner",    "/become-a-partner",       "footer link", "Company"),
    ("footer", "CloudEdge",           "/cloudedge",              "footer link", "Solutions"),
    ("footer", "Posta",               "/posta",                  "footer link", "Solutions"),
    ("footer", "CloudSpace",          "/cloudspace",             "footer link", "Solutions"),
    ("footer", "Knowledge Base",      "/knowledge-base",         "footer link", "Resources"),
    ("footer", "Privacy Policy",      "/privacy-policy",         "footer link", "Resources"),
    ("footer", "Terms of Service",    "/terms",                  "footer link", "Resources"),
    ("footer", "Contact Us",          "/contact-us",             "footer link", "Resources"),
    ("footer", "Facebook",            "https://facebook.com/worldposta",                 "social link", "Social"),
    ("footer", "LinkedIn",            "https://www.linkedin.com/company/worldposta/",    "social link", "Social"),
    ("footer", "Twitter/X",           "https://twitter.com/worldposta",                  "social link", "Social"),
]

SKIP_DEST = {"sso-dynamic", "form-submit", "calendly-popup", "js-function"}
SKIP_LABELS = {
    "sso-dynamic":    "⏭ SSO dynamic — shows 'My Dashboard' when logged in (not testable unauthenticated)",
    "form-submit":    "⏭ Form submit button — no navigation target",
    "calendly-popup": "⏭ Opens Calendly popup via JS (no URL to check)",
    "js-function":    "⏭ Triggers JS function — no navigation URL",
}


# ─────────────────────────────────────────────────────────────────────────────
# CHROME DRIVER  (identical setup to working registration script)
# ─────────────────────────────────────────────────────────────────────────────

def get_chrome_major() -> int | None:
    try:
        out = subprocess.check_output(["/usr/bin/google-chrome", "--version"]).decode()
        m   = re.search(r"(\d+)\.", out)
        return int(m.group(1)) if m else None
    except Exception:
        return None


def make_driver(headless: bool = True):
    print("🌐 Launching Chrome...")
    os.makedirs(SCREENSHOT_DIR, exist_ok=True)

    options = uc.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-extensions")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    if headless:
        options.add_argument("--headless=new")

    chrome_major = get_chrome_major()
    print(f"   Detected Chrome major: {chrome_major}")

    driver = uc.Chrome(
        options=options,
        browser_executable_path="/usr/bin/google-chrome",
        version_main=chrome_major,
        driver_executable_path=None,
        use_subprocess=True,
    )
    driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
    print("✅ Chrome launched")
    return driver


# ─────────────────────────────────────────────────────────────────────────────
# HTTP CHECKER  (requests — no browser needed)
# ─────────────────────────────────────────────────────────────────────────────

_session = requests.Session()
_session.headers.update(HEADERS)
_session.max_redirects = 10
_http_cache: dict[str, tuple] = {}


def check_url(url: str) -> tuple:
    """Returns (final_url, status_code, error_msg). Cached per URL."""
    if url in _http_cache:
        return _http_cache[url]
    try:
        r = _session.head(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        if r.status_code in (405, 501, 403):
            r = _session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True, stream=True)
        result = (r.url, r.status_code, None)
    except requests.exceptions.SSLError as e:
        result = (url, None, f"SSL Error: {str(e)[:90]}")
    except requests.exceptions.ConnectionError as e:
        result = (url, None, f"Connection Error: {str(e)[:90]}")
    except requests.exceptions.Timeout:
        result = (url, None, f"Timeout after {REQUEST_TIMEOUT}s")
    except Exception as e:
        result = (url, None, f"Error: {str(e)[:90]}")
    _http_cache[url] = result
    return result


def resolve(path: str) -> str:
    if path.startswith("http"):
        return path
    return BASE_URL.rstrip("/") + "/" + path.lstrip("/")


def build_issues(code, final_url, error) -> str | None:
    if error:               return f"❌ {error}"
    if code == 404:         return f"❌ 404 Not Found"
    if code == 403:         return f"⚠️ 403 Forbidden"
    if code and code >= 500: return f"❌ Server Error {code}"
    if code and 300 <= code < 400: return f"↪ Redirects → {final_url}"
    return None


def categorise(code, issues: str) -> str:
    if not issues or issues == "✅ Clean": return "OK"
    if "⏭" in issues:   return "SKIPPED"
    if "❌" in issues:
        if "404" in issues:    return "404 NOT FOUND"
        if "Timeout" in issues: return "TIMEOUT"
        return "ERROR"
    if "⚠️" in issues:   return "WARNING"
    if "↪" in issues:    return "REDIRECT"
    return "UNKNOWN"


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 1 — PAGE HEALTH  (pure HTTP, no browser)
# ─────────────────────────────────────────────────────────────────────────────

def check_all_pages() -> list[dict]:
    print(f"\n{'='*60}")
    print(f"PHASE 1 — Page Health ({len(SITEMAP_PAGES)} pages)")
    print(f"{'='*60}")

    results = []
    for name, path in SITEMAP_PAGES:
        url             = resolve(path)
        final, code, err = check_url(url)
        issues          = build_issues(code, final, err) or "✅ Clean"
        cat             = categorise(code, issues)

        marker = "✅" if issues == "✅ Clean" else ("↪" if "↪" in issues else "❌")
        print(f"  {marker} [{str(code):>3}] {name:40s} {path}")

        results.append({
            "page_name":       name,
            "path":            path,
            "full_url":        url,
            "final_url":       final,
            "status_code":     code,
            "status_category": cat,
            "issues":          issues,
            "timestamp":       datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
        time.sleep(0.2)

    return results


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 2 — CTA BUTTONS  (HTTP dest check + Selenium DOM presence check)
# ─────────────────────────────────────────────────────────────────────────────

def check_all_ctas(headless: bool) -> list[dict]:
    print(f"\n{'='*60}")
    print(f"PHASE 2 — CTA Buttons ({len(CTA_MAP)} buttons)")
    print(f"{'='*60}")

    results     = []
    dom_cache   = {}      # source_url → page_source string
    driver      = None

    # Identify pages that need DOM check (exclude nav/footer abstract rows)
    pages_needing_dom = {
        resolve(src) for src, *_ in CTA_MAP
        if src not in ("nav", "footer") and _[1] not in SKIP_DEST
    }

    if pages_needing_dom:
        driver = make_driver(headless)

    try:
        # Pre-load DOM for pages that need button presence checks
        if driver:
            print(f"\n  Loading {len(pages_needing_dom)} pages for DOM checks...")
            for url in sorted(pages_needing_dom):
                if url in dom_cache:
                    continue
                try:
                    driver.get(url)
                    time.sleep(3)                   # let JS render
                    dom_cache[url] = driver.page_source.lower()
                    print(f"    ✅ Loaded: {url}")
                except Exception as e:
                    dom_cache[url] = ""
                    print(f"    ⚠ Could not load {url}: {e}")

        # Now evaluate every CTA
        for source, text, dest, css_hint, section in CTA_MAP:

            # ── Non-navigating: skip HTTP check, label clearly ─────────────────
            if dest in SKIP_DEST:
                results.append(_cta_row(
                    source, text, dest, css_hint, section,
                    expected_url=dest, final_url=dest,
                    status_code=None, issues=SKIP_LABELS[dest],
                ))
                print(f"  ⏭  SKIP  [{text[:35]:35s}]")
                continue

            # ── HTTP-check the expected destination ───────────────────────────
            expected_full   = resolve(dest)
            final, code, err = check_url(expected_full)
            issues          = build_issues(code, final, err)

            # ── DOM presence check (does button text appear on its page?) ──────
            dom_note = ""
            if source not in ("nav", "footer"):
                src_url = resolve(source)
                page_src = dom_cache.get(src_url, "")
                if page_src and text.lower() not in page_src:
                    dom_note = f" | ⚠️ Button text '{text}' not found in rendered DOM"

            final_issues = (issues or "✅ Clean") + dom_note

            marker = "✅" if not issues else ("↪" if issues and "↪" in issues else "❌")
            print(f"  {marker} [{str(code):>3}] [{text[:30]:30s}] {source:35s} → {dest}")

            results.append(_cta_row(
                source, text, dest, css_hint, section,
                expected_url=expected_full,
                final_url=final,
                status_code=code,
                issues=final_issues,
            ))
            time.sleep(0.1)

    finally:
        if driver:
            try:
                driver.quit()
                print("\n✅ Browser closed")
            except Exception:
                pass

    return results


def _cta_row(source, text, dest, css_hint, section,
             expected_url, final_url, status_code, issues) -> dict:
    return {
        "source_page":     source,
        "section":         section,
        "cta_text":        text,
        "expected_dest":   dest,
        "css_hint":        css_hint,
        "expected_url":    expected_url,
        "final_url":       final_url or "",
        "status_code":     status_code,
        "status_category": categorise(status_code, issues),
        "issues":          issues,
        "timestamp":       datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL REPORT
# ─────────────────────────────────────────────────────────────────────────────

C_NAV  = "1E3A5F"
C_OK_A = "D4EDDA"
C_OK_B = "F0FFF4"
C_ERR  = "F8D7DA"
C_WARN = "FFF3CD"
C_SKIP = "EEF2FF"
C_SUM  = "EBF0F7"


def _fill(c):
    return PatternFill("solid", fgColor=c)

def _font(bold=False, size=10, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)

def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def _header_row(ws, cols, bg=C_NAV):
    for i, (title, width) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=title)
        c.fill      = _fill(bg)
        c.font      = _font(bold=True, size=11, color="FFFFFF")
        c.alignment = _align("center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

def _row_bg(issues: str) -> str:
    if "⏭" in issues: return C_SKIP
    if "❌" in issues: return C_ERR
    if "⚠️" in issues: return C_WARN
    if "↪"  in issues: return C_WARN
    return None


def generate_excel(page_results: list, cta_results: list, path: str):
    wb = openpyxl.Workbook()

    # ── Sheet 1 — Page Health ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Page Health"
    cols1 = [
        ("Page Name", 30), ("Path", 38), ("Full URL", 48),
        ("Final URL", 48), ("HTTP Status", 12), ("Category", 18), ("Issues / Notes", 55),
    ]
    _header_row(ws1, cols1)

    for i, r in enumerate(page_results, 2):
        bg = _row_bg(r["issues"]) or (C_OK_A if i % 2 == 0 else C_OK_B)
        for col, val in enumerate(
            [r["page_name"], r["path"], r["full_url"], r["final_url"],
             r["status_code"], r["status_category"], r["issues"]], 1
        ):
            c = ws1.cell(row=i, column=col, value=val)
            c.fill = _fill(bg)
            c.font = _font(size=9 if col in (3, 4) else 10)
            c.alignment = _align(wrap=(col == 7))
        ws1.row_dimensions[i].height = 18

    # ── Sheet 2 — CTA Buttons ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("CTA Buttons")
    cols2 = [
        ("Source Page", 28), ("Section", 28), ("CTA Text", 32),
        ("Expected Destination", 38), ("CSS Classes", 38),
        ("HTTP Status", 12), ("Category", 18), ("Issues / Notes", 60),
    ]
    _header_row(ws2, cols2)

    for i, r in enumerate(cta_results, 2):
        bg = _row_bg(r["issues"]) or (C_OK_A if i % 2 == 0 else C_OK_B)
        for col, val in enumerate(
            [r["source_page"], r["section"], r["cta_text"],
             r["expected_dest"], r["css_hint"],
             r["status_code"], r["status_category"], r["issues"]], 1
        ):
            c = ws2.cell(row=i, column=col, value=val)
            c.fill = _fill(bg)
            c.font = _font(size=9 if col in (4, 5) else 10)
            c.alignment = _align(wrap=(col == 8))
        ws2.row_dimensions[i].height = 18

    # ── Sheet 3 — Issues Only ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("⚠ Issues Only")
    cols3 = [
        ("Type", 10), ("Source", 30), ("Element / Page", 30),
        ("Expected Dest", 38), ("HTTP Status", 12), ("Category", 18), ("Issues", 60),
    ]
    _header_row(ws3, cols3, bg="C0392B")
    row3 = 2

    for r in page_results:
        if "❌" not in r["issues"] and "⚠️" not in r["issues"]:
            continue
        for col, val in enumerate(
            ["PAGE", r["path"], r["page_name"], r["full_url"],
             r["status_code"], r["status_category"], r["issues"]], 1
        ):
            c = ws3.cell(row=row3, column=col, value=val)
            c.fill = _fill(C_ERR)
            c.font = _font(size=10)
            c.alignment = _align(wrap=(col == 7))
        ws3.row_dimensions[row3].height = 18
        row3 += 1

    for r in cta_results:
        if "❌" not in r["issues"] and "⚠️" not in r["issues"]:
            continue
        for col, val in enumerate(
            ["CTA", r["source_page"], r["cta_text"], r["expected_dest"],
             r["status_code"], r["status_category"], r["issues"]], 1
        ):
            c = ws3.cell(row=row3, column=col, value=val)
            c.fill = _fill(C_ERR if "❌" in r["issues"] else C_WARN)
            c.font = _font(size=10)
            c.alignment = _align(wrap=(col == 7))
        ws3.row_dimensions[row3].height = 18
        row3 += 1

    if row3 == 2:
        c = ws3.cell(row=2, column=1, value="🎉 No issues found — all pages and CTAs are healthy!")
        c.fill = _fill(C_OK_A)
        c.font = _font(bold=True, size=12)

    # ── Sheet 4 — Summary ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("📊 Summary")
    ws4.column_dimensions["A"].width = 38
    ws4.column_dimensions["B"].width = 18

    p_ok    = sum(1 for r in page_results if r["status_category"] == "OK")
    p_redir = sum(1 for r in page_results if r["status_category"] == "REDIRECT")
    p_warn  = sum(1 for r in page_results if "⚠️" in r["issues"])
    p_err   = sum(1 for r in page_results if "❌" in r["issues"])
    c_ok    = sum(1 for r in cta_results  if r["status_category"] == "OK")
    c_skip  = sum(1 for r in cta_results  if r["status_category"] == "SKIPPED")
    c_redir = sum(1 for r in cta_results  if r["status_category"] == "REDIRECT")
    c_err   = sum(1 for r in cta_results  if "❌" in r["issues"])
    total   = p_err + c_err + p_warn

    rows = [
        ("WorldPosta Link & CTA Check Report", ""),
        ("Generated At",  datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Target",        BASE_URL),
        ("", ""),
        ("── PAGE HEALTH ──", ""),
        ("Total Pages (from sitemap)",  len(page_results)),
        ("✅ OK (200)",                  p_ok),
        ("↪  Redirects (3xx)",          p_redir),
        ("⚠️  Warnings (403 etc)",       p_warn),
        ("❌ Errors / 404s",             p_err),
        ("", ""),
        ("── CTA BUTTONS ──", ""),
        ("Total CTAs (from mapping)",   len(cta_results)),
        ("✅ Destination OK",            c_ok),
        ("⏭  Skipped (SSO/form/JS)",    c_skip),
        ("↪  Redirecting",              c_redir),
        ("❌ Broken destination",        c_err),
        ("", ""),
        ("TOTAL ISSUES", total),
    ]

    for ri, (label, val) in enumerate(rows, 1):
        la = ws4.cell(row=ri, column=1, value=label)
        vl = ws4.cell(row=ri, column=2, value=val)
        is_title   = ri == 1
        is_section = str(label).startswith("──")
        is_total   = label == "TOTAL ISSUES"

        if is_title:
            la.font = _font(bold=True, size=15, color=C_NAV)
        elif is_section:
            la.fill = _fill(C_NAV); vl.fill = _fill(C_NAV)
            la.font = _font(bold=True, color="FFFFFF", size=11)
        elif is_total:
            bg = "F8D7DA" if total > 0 else C_OK_A
            la.fill = _fill(bg); vl.fill = _fill(bg)
            la.font = _font(bold=True, size=12); vl.font = _font(bold=True, size=12)
        else:
            la.fill = _fill(C_SUM); vl.fill = _fill(C_SUM)
            la.font = _font(size=11); vl.font = _font(size=11)

        la.alignment = _align()
        vl.alignment = _align("center")
        ws4.row_dimensions[ri].height = 24

    wb.save(path)
    print(f"\n  ✅ Excel report saved → {path}")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    global EXCEL_OUTPUT

    parser = argparse.ArgumentParser(description="WorldPosta Precision Link Checker")
    parser.add_argument("--headless", action="store_true",
                        help="Run Chrome headless (required for GitHub Actions)")
    parser.add_argument("--no-fail",  action="store_true",
                        help="Always exit 0 — let n8n parse JSON for pass/fail logic")
    parser.add_argument("--output",   default=EXCEL_OUTPUT,
                        help="Excel output filename")
    args = parser.parse_args()

    EXCEL_OUTPUT = args.output

    os.makedirs(SCREENSHOT_DIR, exist_ok=True)

    print("=" * 60)
    print("🔍  WORLDPOSTA PRECISION LINK & CTA CHECKER")
    print("=" * 60)
    print(f"  Pages   : {len(SITEMAP_PAGES)}")
    print(f"  CTAs    : {len(CTA_MAP)}")
    print(f"  Started : {datetime.now():%Y-%m-%d %H:%M:%S}")

    page_results = check_all_pages()
    cta_results  = check_all_ctas(headless=args.headless)

    # Save JSON
    with open(JSON_OUTPUT, "w", encoding="utf-8") as f:
        json.dump({"pages": page_results, "ctas": cta_results}, f, indent=2, ensure_ascii=False)

    # Save Excel
    generate_excel(page_results, cta_results, EXCEL_OUTPUT)

    # Final summary
    p_err = sum(1 for r in page_results if "❌" in r["issues"])
    c_err = sum(1 for r in cta_results  if "❌" in r["issues"])
    total = p_err + c_err

    print("\n" + "=" * 60)
    print("📊  FINAL SUMMARY")
    print("=" * 60)
    print(f"  Pages checked : {len(page_results)}  |  Issues: {p_err}")
    print(f"  CTAs checked  : {len(cta_results)}   |  Issues: {c_err}")
    print(f"  Total Issues  : {total}")
    print(f"\n  📄 Excel → {EXCEL_OUTPUT}")
    print(f"  📋 JSON  → {JSON_OUTPUT}")

    if total:
        print(f"\n::warning::Link checker found {total} issue(s). See {EXCEL_OUTPUT}")

    # Compact JSON block for n8n to parse from Actions logs
    print("\n===== BEGIN_LINK_CHECK_JSON =====")
    print(json.dumps({
        "total_pages":  len(page_results),
        "total_ctas":   len(cta_results),
        "page_errors":  p_err,
        "cta_errors":   c_err,
        "total_issues": total,
        "report_file":  EXCEL_OUTPUT,
        "timestamp":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }, indent=2))
    print("===== END_LINK_CHECK_JSON =====")

    sys.exit(0 if (args.no_fail or total == 0) else 1)


if __name__ == "__main__":
    main()